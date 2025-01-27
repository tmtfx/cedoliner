#    Cedoliner, elaboration, extraction and computation of pay slip parameters
#    Copyright (C) 2025  Fabio Tomat
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    You should have received a copy of the GNU General Public License
#    along with this program.  If not, see <https://www.gnu.org/licenses/>.

import pdfplumber
import os
import openpyxl
from openpyxl.styles import PatternFill
#from difflib import SequenceMatcher

red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

start_row_for_year = None

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Risultati"
# Aggiungi intestazioni
ws.append(["Mese", "Pagina", "Codice", "Descrizione", "Importo"])

# Specifica la cartella contenente i PDF
cartella_pdf = "cedolini"

# Parole chiave o pattern da cercare
parole_chiave = ["0169", "0170", "0964", "0965", "0966", "0967", "0968", "0987", "0988", "0991", "0992", "0790", "0791"]
mese_anno_ref = ["Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno", "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"]#, "Tredicesima", "Quattordicesima"]

def mese_a_numero(mese):
    mesi = {
        "Gennaio": 1, "Febbraio": 2, "Marzo": 3, "Aprile": 4,
        "Maggio": 5, "Giugno": 6, "Luglio": 7, "Agosto": 8,
        "Settembre": 9, "Ottobre": 10, "Novembre": 11, "Dicembre": 12
    }
    return mesi.get(mese, 0)

def log(testo, azzera_file=False):
    try:
        modalita = "w" if azzera_file else "a"
        with open("log.txt", modalita) as file_log:
            file_log.write(testo + "\n")
    except Exception as e:
        print(f"Errore nella scrittura del file di log: {e}")
#def best_match(stringa, lista_stringhe):
#    miglior_match = None
#    massimo_ratio = 0
#    
#    for elemento in lista_stringhe:
#        ratio = SequenceMatcher(None, stringa, elemento).ratio()
#        if ratio > massimo_ratio:
#            massimo_ratio = ratio
#            miglior_match = elemento
#    
#    return miglior_match, massimo_ratio
def deduci_mese_da_nome_file(pdf_path,anno,isnoloop):
    tn=os.path.splitext(os.path.basename(pdf_path))[0]
    maybe_mese=tn.split()
    no_ref=False
    mese=None
    for mnth in mese_anno_ref:
        if mnth.lower() in [mese.lower() for mese in maybe_mese]:
            mese=mnth
            print("rilevato:",mese,anno)
            break
    if mese is None:
        stripname=tn.replace(" ","")
        stripname=stripname.replace(anno,"")
        p=stripname.lower().find("cedolino")
        if p>-1:
            stripname=stripname[p+len("cedolino"):]
        stripname=stripname.replace("-","")
        stripname=stripname.replace("_","")
        p=stripname.find(".")
        if p>-1:
            stripname=stripname[:p]
        solo_mese = "".join(c for c in stripname if c.isdigit())
        if solo_mese!="":
            try:
                mese=mese_anno_ref[int(solo_mese)-1]
                print("dedotto:",mese,anno)
            except:
                no_ref=True
        else:
            no_ref=True
    if no_ref:
        if isnoloop and mese is None:
            print(f"ATTENZIONE: impossibile rilevare il mese dal nome del file,\nverificare il nome del file per {pdf_path}")
    return mese

def analizza_cedolino(pdf_path, anno, parole_chiave):
    risultati = []
    got_ref=False
    mese = None
    mese_rilevato=None
    ispdf=True
    """questa parte ricerca il mese e l'anno del cedolino"""
    #print("analisi del cedolino",pdf_path)
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page_num, page in enumerate(pdf.pages, start=1):
                testo = page.extract_text()
                if testo:
                    righe = testo.split("\n")
                    for riga in righe:
                        if not got_ref:
                            for parola in mese_anno_ref:
                                if parola in riga:
                                    n=riga.find(parola)
                                    if n!=-1:
                                        mese_anno=riga[n:]
                                        if (mese_anno.split()[1]==anno)or(mese_anno[-4:]==anno):
                                            n=len(parola)
                                            mese_rilevato=mese_anno[:n]
                                            if pdf_path.lower().find(mese_rilevato.lower())>-1:
                                                #anno=mese_anno[n:]
                                                mese=mese_rilevato
                                                print("elaborazione di",mese,anno)
                                                got_ref=True
                                                break
                                            else:
                                                mese=deduci_mese_da_nome_file(pdf_path,anno,False)
                                                if mese is not None:
                                                    got_ref=True
                                                    break

    except Exception as e:
            print(f"Errore nell'analisi del cedolino {pdf_path}: {e}")
            ispdf=False
    
    if ispdf:
        if not got_ref:
            print("fallback - nome mese da nome file")
            deduci_mese_da_nome_file(pdf_path,anno,True)

        """questa parte cerca i codici e le colonne in cui son scritti"""
        
        with pdfplumber.open(pdf_path) as pdf:
            totcoddesc=[]
            """ vecchio ciclo separato integrato in quello successivo
            for page_num, page in enumerate(pdf.pages, start=1):
                tables = page.extract_tables()
                stopiter=False
                lastcode=""
                #for table in tables:
                for table_idx, table in enumerate(tables, start=1):
                    if table:
                        if not stopiter:  # Se la tabella non è vuota e non è già stata trovata
                            for riga in table:
                                #print(f"riga di tabella: {riga}")
                                coddesc=[]
                                #qualche cedolino è formattato diversamente, quindi devo rimuovere i valori nulli
                                newriga=[]
                                for item in riga:
                                    if item is not None:
                                        newriga.append(item)
                                #print(f"newriga: {newriga}")
                                #if riga[1]!=None:
                                try:
                                    elementir0=newriga[0].split("\n")
                                    elementir1=newriga[1].split("\n")
                                    #controlliamo che il numero di elementi della prima e della seconda colonna siano uguali
                                    #codice con loro descrizione
                                    if len(elementir0)==len(elementir1):
                                        i=0
                                        while i < len(elementir0):
                                            if elementir0[i] in parole_chiave:
                                                coddesc.append((elementir0[i],elementir1[i]))
                                                stopiter=True#non serve continuare a cercare in altre tabelle
                                                lastcode=elementir0[len(elementir0)-1]
                                            i+=1
                                        if coddesc!=[]:
                                            ##print(f"coddesc {coddesc}")
                                            totcoddesc.extend(coddesc)
                                except AttributeError:
                                    #di solito è quando trova un None, quindi non è un problema 
                                    #anzi non dovrebbe più trovare None visto che li abbiamo rimossi
                                    pass
                                except IndexError:
                                    #di solito rileva un codice altrove, quindi non è un problema
                                    pass
                if lastcode == "":
                    print(f"ATTENZIONE: nessun ultimo codice rilevato nella pagina {page_num} del cedolino {pdf_path}")"""
            #print(f"totcoddesc {totcoddesc}")
            for page_num, page in enumerate(pdf.pages, start=1):
                testo = page.extract_text()
                tables = page.extract_tables()
                stopiter=False
                lastcode=""
                #non serve enumerare le tabelle, basta che siano presenti
                #for table_idx, table in enumerate(tables, start=1):
                for table in tables:
                    if table:  # Se la tabella non è vuota
                        #print(f"Pagina {page_num}, Tabella {table_idx}:\n")
                        if not stopiter:  # Se la tabella non è vuota e non è già stata trovata
                            for riga in table:
                                #print(f"riga di tabella: {riga}")
                                coddesc=[]
                                #qualche cedolino è formattato diversamente, quindi devo rimuovere i valori nulli
                                newriga=[]
                                for item in riga:
                                    if item is not None:
                                        newriga.append(item)
                                #print(f"newriga: {newriga}")
                                #if riga[1]!=None:
                                try:
                                    elementir0=newriga[0].split("\n")
                                    elementir1=newriga[1].split("\n")
                                    #controlliamo che il numero di elementi della prima e della seconda colonna siano uguali
                                    #codice con loro descrizione
                                    if len(elementir0)==len(elementir1):
                                        i=0
                                        while i < len(elementir0):
                                            if elementir0[i] in parole_chiave:
                                                coddesc.append((elementir0[i],elementir1[i]))
                                                stopiter=True#non serve continuare a cercare in altre tabelle
                                                lastcode=elementir0[len(elementir0)-1]
                                            i+=1
                                        if coddesc!=[]:
                                            ##print(f"coddesc {coddesc}")
                                            totcoddesc.extend(coddesc)
                                except AttributeError:
                                    #di solito è quando trova un None, quindi non è un problema 
                                    #anzi non dovrebbe più trovare None visto che li abbiamo rimossi
                                    pass
                                except IndexError:
                                    #di solito rileva un codice altrove, quindi non è un problema
                                    pass
                        if page_num==1:
                            colonne = list(zip(*table))
                            for col_idx, colonna in enumerate(colonne, start=1):
                                listacolonna=[]
                                for item in colonna:
                                    if item is not None:
                                        listacolonna.append(item)
                                # uncomment these lines if needed
                                #if listacolonna[0]=="Trattenute":
                                #    trattenute=listacolonna[1].split("\n")
                                #    print(f"Colonna {col_idx} Trattenute: {trattenute}\n")
                                #elif listacolonna[0]=="Competenze":
                                #    competenze=listacolonna[1].split("\n")
                                #    print(f"Colonna {col_idx} Competenze: {competenze}\n")
                                #if listacolonna[0]=="Descrizione":
                                #    #print(f"Colonna {col_idx} Descrizione: {listacolonna}\n")
                                #    descrizioni=listacolonna[1].split("\n")       
                                if col_idx==1:
                                    i=0
                                    while i < len(listacolonna):
                                        if listacolonna[i]=="Presenze":
                                            if listacolonna[i+1]:
                                                if listacolonna[i+1].isnumeric():
                                                    presenze=listacolonna[i+1]
                                                    #print(f"Presenze: {presenze}\n")
                                                else:
                                                    presenze=listacolonna[i+1]
                                                    #print(f"Presenze formato testo: {presenze}\n")
                                            else:
                                                #print("Nessuna presenza\n")
                                                presenze="0"
                                        i+=1
                                else:
                                    if listacolonna[0]=="Ferie":
                                        if listacolonna[1]:
                                            if listacolonna[1].isnumeric():
                                                ferie=listacolonna[1]
                                                #print(f"Ferie: {ferie}\n")
                                            else:
                                                ferie=listacolonna[1]
                                                #print(f"Ferie formato testo: {ferie}\n")
                                        else:
                                            #print("Nessun giorno di ferie\n")
                                            ferie="0"
                if lastcode == "":
                    #pdf_path
                    log(f"la pagina {page_num} del cedolino [{os.path.basename(pdf_path)}] nella cartella \"{anno}\" potrebbe aver prodotto risultati inattesi, verificare")
                    print(f"ATTENZIONE: nessun ultimo codice rilevato nella pagina {page_num} del cedolino {pdf_path}")                           
                if testo:
                    startelaborate=False
                    elaborateandquit=False
                    righe = testo.split("\n")
                    #print(f"Testo pagina {page_num}:\n{righe}")
                    for numr,riga in enumerate(righe):
                        if not startelaborate:
                            if riga.lower().replace(" ","").find("cod.voce")>-1:
                                startelaborate=True
                                continue
                            else:
                                continue
                        if startelaborate:
                            if lastcode!="":
                                if lastcode in riga:
                                    #print(f"Riga con lastcode {lastcode}: {riga}")
                                    try:
                                        if lastcode in righe[numr+1]:
                                            #print(f"Riga successiva con lastcode {lastcode}: {righe[numr+1]}")
                                            pass
                                        else:
                                            elaborateandquit=True #non serve cercare oltre
                                    except IndexError:
                                        elaborateandquit=True #non serve cercare oltre
                            #print(f"riga di testo: {riga}")
                            #if any(parola in riga for parola in parole_chiave):
                            for parola in parole_chiave:
                                #if parola.lower() in riga.lower(): sono numeri non serve il lower
                                if parola in riga:
                                    #print(f"Parola chiave {parola} trovata in riga: {riga}")
                                    words=riga.split()
                                    if parola in words:

                                #decommenta per stampare a schermo le righe estratte
                                #print(f"Riga: {riga}")
                                #for parola in parole_chiave:
                                    #if parola.lower() in riga.lower():
                                        #recupero descrizione
                                        
                                        #questa parte serve a evitare che il codice venga rilevato in alto a dx
                                        #TODO: forse è meglio usare l'analisi della tabella per evitare questo problema
                                        #ma per ora lascio così
                                        ns=riga.lower().find(parola.lower())+len(parola)
                                        ne=riga.find(" X ")
                                        if ne==-1:
                                            ne=len(riga)-len(parola)
                                        descrizione=riga[ns:ne]
                                        #print(f"descrizione: {descrizione}")
                                        if descrizione!="":
                                            #best_match fornisce una stima alcune volte sbagliata, 
                                            #meglio evitare anche se il testo è molto più gradevole
                                            #EDIT: ho trovato un metodo migliore, questa è l'implementazione:
                                            #      nel caso dia problemi è possibile rimuovere questo metodo
                                            #      e usare la stringa sopra, sebbene talvolta presenti valori
                                            #      aggiuntivi non inerenti
                                            #metodo esplicito
                                            #print("descrizione prima: ",descrizione)
                                            if parola in [x[0] for x in totcoddesc]:
                                                for x in totcoddesc:
                                                    if x[0]==parola:
                                                        descrizione=x[1]
                                                        totcoddesc.remove(x)
                                                        #print(f"abbellimento: {descrizione}")
                                                        break
                                            #print("descrizione dopo: ",descrizione)
                                            #promemoria programmazione:
                                            #metodo comprensione liste
                                            #descrizione=[x[1] for x in totcoddesc if x[0]==parola][0]
                                            #metodo più sicuro con comprensione liste
                                            #descrizione = next((x[1] for x in totcoddesc if x[0] == parola), None)

                                            valori=riga.split()
                                            if valori[-2].find("-")>-1:
                                                # trovata trattenuta mediante aliquota o parametro negativo
                                                valore="-"+valori[-1]
                                            else:
                                                # trovata competenza mediante aliquota o parametro positivo
                                                valore=valori[-1]
                                            risultati.append((page_num, mese, parola, valore,descrizione))
                        if elaborateandquit:
                            break
                    
                    if lastcode == "":
                        tst=""
                        for it in risultati:
                            if it[0]==page_num:
                                tst+=f"{it}\n"
                        log(f"valori della pagina {page_num} inseriti:\n{tst}\n")
                #    print(f"ATTENZIONE: nessun ultimo codice rilevato nella pagina {page_num} del cedolino {pdf_path}")
        if len(totcoddesc)>0:
            log(f"il file [{pdf_path}] nella cartella \"{anno}\" ha codici non assegnati, ne mancano: {len(totcoddesc)}\npiù precisamente mancano: {totcoddesc}\n")
            print(f"ATTENZIONE: qualche descrizione in questo mese potrebbe essere sbagliata\nmancano da assegnare: {totcoddesc}")
        return (risultati,(presenze,ferie,mese))
    else:
        return (None,(None,None,None))


# Ottieni la lista di tutti i file PDF nella cartella
log("Registro errori e avvisi",True)
log("________________________\n")
pdf_files = []
risultati_cartella = []
for root, dirs, files in os.walk(cartella_pdf):
    for file in files:
        if file.lower().endswith('.pdf'):
            pdf_files.append((os.path.join(root, file), os.path.basename(root)))
    if os.path.basename(root).isnumeric():
        ws.append(["Anno:", os.path.basename(root)])
        cell = ws.cell(row=ws.max_row, column=1)
        cell.font = openpyxl.styles.Font(size=14, bold=True)
        cell = ws.cell(row=ws.max_row, column=2)
        cell.font = openpyxl.styles.Font(size=14, bold=True)
    conta=0
    pf_risultati = []
    for pdf_path in pdf_files:
        conta+=1
        ret = analizza_cedolino(pdf_path[0], pdf_path[1], parole_chiave)#, pattern_codici)
        if ret == (None, (None, None, None)):
            print(f"ATTENZIONE: l'analisi di [{pdf_path[0]}] non ha prodotto risultati")
            #pdf_path[0]
            log(f"l'analisi di [{os.path.basename(pdf_path[0])}] nella cartella \"{os.path.basename(root)}\" non ha prodotto risultati, probabilmente il file è rovinato\n")
            continue
        risultati,(pres,fer,month) = ret
        if month is not None or pres is not None or fer is not None:
            pf_risultati.append((month,pres,fer))
        else:
            #pdf_path[0]
            log(f"nel file [{os.path.basename(pdf_path[0])}] della cartella \"{os.path.basename(root)}\" non è stato possibile rilevare il mese o le presenze o le ferie\n")
            print(f"ATTENZIONE: nel file {pdf_path[0]} non è stato possibile rilevare il mese o le presenze o le ferie")
        #ws.append([f"Presenze: {month}", pres, "Ferie:", fer])
        if risultati is None:
            #pdf_path[0]
            log(f"il file [{os.path.basename(pdf_path)}] nella cartella \"{os.path.basename(root)}\" non ha prodotto risultati, anche se il file è stato letto\n")
            print(f"ATTENZIONE: il file [{pdf_path[0]}] non ha prodotto risultati, anche se il file è stato letto")
        else:
            risultati_cartella.extend(risultati)
    pf_risultati_ordinati = sorted(pf_risultati, key=lambda x: mese_a_numero(x[0]))
    for mese,pres,fer in pf_risultati_ordinati:
        ws.append([f"Presenze {mese}:", pres, f"Ferie {mese}:", fer])
    if conta<12:
        print("ATTENZIONE: sono stati elaborati meno di 12 file")
        log(f"nella cartella \"{os.path.basename(root)}\" sono stati elaborati meno di 12 file\nverifica che ci siano tutti i file\n")
    elif conta==12:
        log(f"nella cartella \"{os.path.basename(root)}\" sono stati elaborati 12 file\nverifica che la tredicesima sia all'interno di uno dei file o che ci siano tutti i file\n")
        print("ATTENZIONE: sono stati elaborati 12 files, verifica che:\nla tredicesima sia segnata all'interno di uno dei file o che ci siano tutti i file")
    else:
        print("file elaborati:",conta)
    risultati_ordinati = sorted(risultati_cartella, key=lambda x: mese_a_numero(x[1]))

    for pagina, mese, parola,valore,descrizione in risultati_ordinati:
        if start_row_for_year is None:
            start_row_for_year = ws.max_row + 1  # Set the starting row for the current year
        ws.append([mese, str(pagina), parola,descrizione, float(valore.replace(",","."))])
        cell = ws.cell(row=ws.max_row,column=5)
        if str(valore).startswith("-"):
            cell.fill = red_fill
    if start_row_for_year is not None:
        ws.append(["","","",f"Totale {pdf_path[1]}", f"=SUM(E{start_row_for_year}:E{ws.max_row})"])
    pdf_files.clear()# = []
    risultati_cartella.clear()  # Clear the results for the next year
    start_row_for_year = None

wb.save("risultati_cedolini.xlsx")

print("File Excel creato con successo: risultati_cedolini.xlsx")
